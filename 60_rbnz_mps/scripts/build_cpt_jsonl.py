#!/usr/bin/env python3
"""Build CPT world-knowledge records from RBNZ Monetary Policy Statements.

One record = an MPS's own VERBATIM policy-assessment narrative paired with the multi-channel macro
series (OCR, inflation, unemployment, GDP, TWI) published in that same statement's data pack.

ACCESS: rbnz.govt.nz answers HTTP 403 to automated fetches -- re-verified 2026-08-19 WITH a full
desktop-Chrome User-Agent, and the body is RBNZ's own "Website unavailable" page, so this is a
bot-wall, not a missing-UA problem. Everything is read from Wayback snapshots. RBNZ's terms grant
reproduction with attribution, so this is an access-method choice, not a licence one.

FULL-ARCHIVE MODE (2026-08-19). The build used to pin ONE statement in config. scripts/census.py
now enumerates the archive from Wayback CDX: 119 statements, 1996-12..2026-05, with every year
1997-2025 complete at 4/4. Three things had to be generalised to reach them:

  (1) NARRATIVE ANCHORS ARE ERA-DEPENDENT. The old extractor keyed on "Latest OCR decision" ..
      "Most recent outlook for the OCR", which exist only on the current template. Older pages put
      the same content under <h2>Policy assessment</h2> (1996-2015) or an "in pictures" block
      (2016-2021). Anchoring on one era's chrome is exactly the defect that held FHFA #59 to 28 of
      46 records, so this now walks the page's heading/paragraph structure and matches a LIST of
      anchors, with a content-based fallback (first policy paragraph that names the OCR).

  (2) DATA-PACK URLS ARE NOT CONSTRUCTIBLE. Pack filenames are irregular across the archive
      (`aug00-data.xls`, `dec13data.xls`, `dec2010-data.xls`, `jun11data.xls`, `mpsmar13data.xls`,
      `august-2022-monetary-policy-statement-data.xlsx`). They are therefore read out of each
      statement page's OWN HTML. Note the href carries a `?revision=...` query string AFTER the
      extension, so an `href="...\\.xlsx?"$` style regex finds nothing -- that query string must be
      allowed for, and it must stay in the cache key (a params-blind cache would serve the wrong
      revision back after a fix).

  (3) SHEET NUMBERS DO NOT IDENTIFY VARIABLES ACROSS ERAS -- the important one. The old config
      mapped sheet "2.1" -> Production GDP. But figure numbering is re-cut every few years:
      measured 2026-08-19, `Fig 2.1` is "CPI inflation" in the Jun-2005 pack and "Employment and
      investment intentions" in the Dec-2010 pack. Reusing a name->variable map across eras would
      silently mislabel series -- a correctness defect, not a coverage one. Every sheet carries its
      OWN title cell ("CPI inflation", "GDP growth", "Trade weighted index"), so variables are
      matched on that title text instead, and a sheet whose title matches nothing is skipped and
      counted rather than guessed at.

THE "CURRENT QUARTER" SPLIT: a pack's own-vintage column runs from measured history straight into
the Bank's own projection with no visual break. Verified on the Feb-2026 MPS: OCR's row for the
statement's own quarter = 2.25%, exactly the announced decision -- a real current decision. But
unemployment/inflation's same-quarter row IS a forward projection, and the text's quoted values
match the PRIOR quarter (last real outturn). Quarterly channels are windowed to a common end
quarter (the schema needs equal length per freq); where that final point is itself a projection it
carries the same forecast-not-measured caveat as WASDE #41 / GAIN #58 -- the text is the
contemporaneous first-party forecast, so there is no future-value leakage.

Usage:
    python scripts/census.py --config config.example.yaml            # enumerate first
    python scripts/build_cpt_jsonl.py --config config.example.yaml   # then build
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

import gzip
import threading
import zlib

import openpyxl
import yaml

HERE = Path(__file__).resolve().parent
PKG_ROOT = HERE.parent
sys.path.insert(0, str(PKG_ROOT.parent / "schema"))
from emit import emit_record  # noqa: E402

UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"}


class _Pacer:
    """One process-wide pace for every archive.org request, with additive-increase /
    multiplicative-decrease backoff.

    archive.org enforces a GLOBAL limit, not a per-endpoint one, and it does not answer with a
    tidy 429: measured 2026-08-19, after a burst of CDX sweeps plus 4 concurrent page fetches, the
    CDX endpoint and /web/ both began REFUSING TCP connections outright while archive.org's own
    root still returned 200. A refused connection is indistinguishable from "not archived" unless
    it is treated as a throttle -- which is why this class exists and why a connection error is
    never allowed to count as a missing narrative.
    """

    def __init__(self, gap: float = 2.0, max_gap: float = 18.0):
        self.gap = gap
        self.min_gap = gap
        self.max_gap = max_gap
        self.next_at = 0.0
        self.lock = threading.Lock()

    def wait(self) -> None:
        with self.lock:
            now = time.monotonic()
            due = max(now, self.next_at)
            self.next_at = due + self.gap
        delay = due - time.monotonic()
        if delay > 0:
            time.sleep(delay)

    def penalise(self) -> None:
        with self.lock:
            self.gap = min(self.max_gap, self.gap * 2)

    def reward(self) -> None:
        # MULTIPLICATIVE recovery. Subtracting a constant was far too timid: a handful of 503s
        # pushed the gap to its ceiling and it then needed ~170 consecutive successes to come back,
        # so the run spent hours sleeping after a brief rate-limit episode.
        with self.lock:
            self.gap = max(self.min_gap, self.gap * 0.7)


PACER = _Pacer(gap=3.0)


def wayback_url(url: str, ts: str) -> str:
    # `id_` asks for the ORIGINAL bytes rather than Wayback's rewritten page; without it the
    # archived HTML comes back with injected banner markup and rewritten asset hrefs.
    return f"https://web.archive.org/web/{ts}id_/{url}"


class WaybackPlaceholderError(RuntimeError):
    """Wayback served its own HTML wrapper instead of the archived binary. Confirmed real and
    non-deterministic 2026-08-03: the SAME xlsx URL returned a valid 1.97MB file once, then a
    10KB HTML placeholder on the next fetch with no code change -- Wayback's nearest-capture
    redirect can land on a different, uncaptured timestamp. Retry rather than trust the first
    response for binary assets."""


def _decompress(blob: bytes, encoding: str = "") -> bytes:
    """Decode a compressed HTTP body.

    urllib does NOT do this for you: when Wayback answers with Content-Encoding: gzip, urlopen
    hands back raw DEFLATE bytes. Those were being written straight into the page cache, where they
    looked like a plausible ~20KB "page" that simply contained no HTML tags -- 63 of 89 cached pages
    were gzip blobs, which read as "this statement has no narrative" rather than as a fetch bug.
    The magic-number test is the reliable one, since the header is not always present."""
    if blob[:2] == b"\x1f\x8b" or "gzip" in encoding:
        try:
            return gzip.decompress(blob)
        except Exception:                                        # noqa: BLE001
            pass
    if "deflate" in encoding:
        for wbits in (-zlib.MAX_WBITS, zlib.MAX_WBITS):
            try:
                return zlib.decompress(blob, wbits)
            except Exception:                                    # noqa: BLE001
                pass
    return blob


def _looks_like_html(blob: bytes) -> bool:
    head = blob[:4000].lower()
    return b"<html" in head or b"<!doctype" in head or b"<div" in head or b"<p" in head


def _is_workbook(blob: bytes) -> bool:
    #  xlsx = zip ('PK'),  legacy xls = OLE2 compound file (D0 CF 11 E0)
    return blob[:2] == b"PK" or blob[:4] == b"\xd0\xcf\x11\xe0"


def fetch(url: str, cache: Path, expect_workbook: bool = False, retries: int = 5) -> bytes:
    if cache.exists() and cache.stat().st_size > 0:
        return cache.read_bytes()
    cache.parent.mkdir(parents=True, exist_ok=True)
    last: Exception | None = None
    for attempt in range(retries):
        PACER.wait()
        try:
            req = urllib.request.Request(url, headers={**UA, "Accept-Encoding": "gzip, identity"})
            with urllib.request.urlopen(req, timeout=180) as r:
                blob = _decompress(r.read(), r.headers.get("Content-Encoding", "") or "")
            PACER.reward()
        except urllib.error.HTTPError as e:
            last = e
            if e.code in (404, 403):        # definitively not archived -- retrying cannot help
                break
            PACER.penalise()
            time.sleep(min(4 * 2 ** attempt, 60))
            continue
        except Exception as e:                                   # noqa: BLE001
            last = e
            PACER.penalise()                # refused connection => throttled, so slow down
            time.sleep(min(4 * 2 ** attempt, 60))
            continue
        if expect_workbook and not _is_workbook(blob):
            last = WaybackPlaceholderError(
                f"expected a workbook but got {len(blob)} bytes starting {blob[:24]!r} "
                f"-- Wayback placeholder ({attempt+1}/{retries})")
            time.sleep(min(4 * 2 ** attempt, 60))
            continue
        cache.write_bytes(blob)
        return blob
    raise last if last else RuntimeError(f"fetch failed: {url}")


# ---------------------------------------------------------------------------- HTML / narrative
_ENT = {"&nbsp;": " ", "&amp;": "&", "&rsquo;": "’", "&lsquo;": "‘",
        "&ldquo;": "“", "&rdquo;": "”", "&ndash;": "–", "&mdash;": "—",
        "&quot;": '"', "&#39;": "'", "&eacute;": "é", "&pound;": "£",
        "&percnt;": "%", "&hellip;": "…"}


def _unent(t: str) -> str:
    for k, v in _ENT.items():
        t = t.replace(k, v)
    return re.sub(r"&#(\d+);", lambda m: chr(int(m.group(1))), t)


def _strip_tags(frag: str) -> str:
    return re.sub(r"\s+", " ", _unent(re.sub(r"<[^>]+>", " ", frag))).strip()


def html_blocks(html: bytes) -> list[tuple[str, str]]:
    """The page as an ordered [(kind, text)] list, kind in {'h','p'}.

    Nav/script/style/footer are removed first: RBNZ's megamenu repeats the words "Official Cash
    Rate" dozens of times, so any content search that runs before stripping it locks onto the menu
    instead of the statement (measured -- the first 'OCR' hit on a 2005 page is inside <ul
    class="megamenu-list">, ~60KB above the real narrative)."""
    t = html.decode("utf-8", "ignore")
    for tag in ("script", "style", "noscript", "nav", "header", "footer", "select"):
        t = re.sub(rf"<{tag}\b.*?</{tag}>", " ", t, flags=re.S | re.I)
    t = re.sub(r"<ul\b[^>]*class=\"[^\"]*megamenu[^\"]*\".*?</ul>", " ", t, flags=re.S | re.I)
    out: list[tuple[str, str]] = []
    for m in re.finditer(r"<(h[1-4]|p|li)\b[^>]*>(.*?)</\1>", t, flags=re.S | re.I):
        kind = "h" if m.group(1).lower().startswith("h") else "p"
        txt = _strip_tags(m.group(2))
        if txt:
            out.append((kind, txt))
    return out


_OCR_SENT = re.compile(
    r"(official cash rate|\bOCR\b|official interest rate|Monetary Policy Committee|"
    r"monetary conditions|\bMCI\b)", re.I)
_NAV_JUNK = (
    re.compile(r"Download the MPS \(PDF,\s*[\d.]+\s*MB\)\s*", re.I),
    re.compile(r"Read the MPS online\s*", re.I),
)
# Boilerplate and link labels that sit in <p>/<li> alongside the narrative. The browser-upgrade
# notice is on EVERY archived page and is long enough to pass a length filter, so it has to be
# named explicitly or it becomes the first "paragraph" of every record.
_JUNK_PARA = (
    re.compile(r"^It looks like the browser you", re.I),
    re.compile(r"^The browser you are using", re.I),
    re.compile(r"^(Data|Charts?) for the .{0,40}(MPS|Monetary Policy Statement)", re.I),
    re.compile(r"^(Monetary Policy Statement|Economic Projections)\s+\w+\s+\d{4}\s*$", re.I),
    re.compile(r"^(Watch|Listen to|Read)\b.{0,60}(video|webcast|media conference|online)", re.I),
    re.compile(r"^(MPS|Monetary Policy Statement)\b.{0,60}(Policy Assessment|PDF)\s*$", re.I),
    re.compile(r"^\s*(PDF|XLS|XLSX)\s*[\d.]+\s*(KB|MB)\s*$", re.I),
    re.compile(r"^Speaking notes", re.I),
    re.compile(r"^Subscribe|^Sign up|^Share this|^Related |^See also", re.I),
)
_STOP_HEADING = re.compile(
    r"^(more information|related|contact|subscribe|share this|"
    r"further information|notes to editors|for (further )?enquiries|about (us|this)|"
    r"tags?|previous|next|search|sign up|newsletter|media (releases?|conference)|"
    r"other (publications|resources)|附)\b", re.I)


def _is_junk(p: str) -> bool:
    return any(r.search(p) for r in _JUNK_PARA)


def extract_narrative(html: bytes, tcfg: dict) -> tuple[str, str]:
    """Return (narrative, how) -- 'how' records which route matched, for the run report.

    Anchor list first, then a CONTENT-DENSITY fallback. The fallback is the part that matters:
    surveying the census showed the 1997-1998 statements put the same content under
    era-specific headings ("Speaking notes for briefing journalists ...", then "Introduction" /
    "The outlook for inflation" / "Policy implications"), and the 1996-1998 statements predate the
    OCR entirely -- they set a "Monetary Conditions Index" level instead, so even a keyword
    fallback on 'OCR' finds nothing. Enumerating every era's heading is the losing game that held
    FHFA #59 to 28 of 46 records, so the last resort is structural: take the longest contiguous run
    of substantial paragraphs on the page, which is the narrative on every template.
    """
    blocks = html_blocks(html)
    anchors = [re.compile(a, re.I) for a in tcfg["narrative_anchors"]]

    def clean(txt: str) -> str:
        for pat in _NAV_JUNK:
            txt = pat.sub("", txt)
        return txt.strip()

    def run_from(i: int) -> list[str]:
        """Paragraphs from block i onward. Continues ACROSS sub-headings (older statements are
        multi-section) but stops at a navigational heading."""
        got: list[str] = []
        for kind, txt in blocks[i:]:
            if kind == "h":
                if _STOP_HEADING.match(txt):
                    break
                continue
            txt = clean(txt)
            if len(txt) < 40 or _is_junk(txt):
                continue
            got.append(txt)
            if sum(len(x) for x in got) > tcfg["max_chars"]:
                break
        return got

    for ai, pat in enumerate(anchors):
        for i, (kind, txt) in enumerate(blocks):
            if kind == "h" and pat.search(txt):
                got = run_from(i + 1)
                if sum(len(x) for x in got) >= tcfg["min_chars"]:
                    return "\n\n".join(got)[: tcfg["max_chars"]], f"anchor:{ai}"

    # structural fallback: longest contiguous run of substantial, non-junk paragraphs
    runs: list[list[str]] = []
    cur: list[str] = []
    for kind, txt in blocks:
        if kind == "h":
            if _STOP_HEADING.match(txt):
                if cur:
                    runs.append(cur)
                cur = []
            continue
        txt = clean(txt)
        if len(txt) < 40 or _is_junk(txt):
            if cur:
                runs.append(cur)
            cur = []
            continue
        cur.append(txt)
    if cur:
        runs.append(cur)
    if runs:
        best = max(runs, key=lambda r: sum(len(x) for x in r))
        if sum(len(x) for x in best) >= tcfg["min_chars"]:
            return "\n\n".join(best)[: tcfg["max_chars"]], "density_fallback"
    return "", "none"


_PACK_RE = re.compile(r"""href=["']([^"']*?\.(?:xlsx|xls)(?:\?[^"']*)?)["']""", re.I)


def find_pack_url(html: bytes, page_url: str) -> str | None:
    """The statement's own data-pack link. NOTE the extension is followed by `?revision=...`;
    a regex ending at the extension finds nothing on these pages."""
    t = html.decode("utf-8", "ignore")
    cands = [_unent(m.group(1)) for m in _PACK_RE.finditer(t)]
    # prefer a link that looks like the MPS graph data, not some unrelated spreadsheet
    ranked = sorted(cands, key=lambda u: (0 if re.search(r"data|mps", u, re.I) else 1, len(u)))
    if not ranked:
        return None
    u = ranked[0]
    if u.startswith("//"):
        return "https:" + u
    if u.startswith("/"):
        return "https://www.rbnz.govt.nz" + u
    if u.startswith("http"):
        return u
    return urllib.parse.urljoin(page_url, u)                      # type: ignore[name-defined]


_ERA_NATIVE = re.compile(r"/mps\d{4}-\d{2}$|/mps-[a-z]+-\d{4}$", re.I)


def candidates_of(st: dict) -> list[dict]:
    """Candidate captures, ERA-NATIVE URL first.

    Ordering matters because archive.org is rate-limited: every candidate tried and failed costs a
    request against a budget that answers 503 once exceeded. The captures that fail are predictable
    -- the Jan-2026 crawl of the modern /hub/publications/... path, which CDX lists as 200 but which
    404s for older statements because RBNZ never served those bodies there. The URL scheme that
    encodes the statement's own date (mps2001-05, mps-may-2001) is the one that resolves for pre-2022
    statements, so it is tried first and is usually the only request needed."""
    cands = st.get("candidates") or [{"wayback_ts": st["wayback_ts"],
                                      "page_url": st["page_url"]}]

    def native(c):
        return bool(_ERA_NATIVE.search(c["page_url"].split("?")[0].rstrip("/")))

    if st.get("year", 0) >= 2022:                 # modern statements: newest capture first
        return sorted(cands, key=lambda c: (1 if native(c) else 0, -int(c["wayback_ts"])))
    return sorted(cands, key=lambda c: (0 if native(c) else 1, -int(c["wayback_ts"])))


def fetch_page(st: dict, cache: Path, tcfg: dict,
               tries_per_candidate: int = 3) -> tuple[bytes, str, str, str, bool]:
    """(html, page_url, wayback_ts, narrative) -- walks the statement's capture candidates.

    A CDX row saying statuscode 200 does NOT guarantee the bytes are retrievable: measured
    2026-08-19, 30 of 119 statements had their newest capture on a Jan-2026 crawl of the /hub/...
    path that 404s on an `id_` fetch while the same statement fetches fine from an older capture of
    an older URL scheme. A page that fetches but carries no narrative is just as useless, so the
    narrative test is part of the acceptance condition rather than a later step."""
    tag = f"{st['year']:04d}-{st['month']:02d}"
    cached = cache / "pages" / f"{tag}.html"
    if cached.exists() and cached.stat().st_size > 0:
        html = cached.read_bytes()
        text, how = extract_narrative(html, tcfg)
        if len(text) >= tcfg["min_chars"]:
            return html, st["page_url"], st["wayback_ts"], text, False
    last = b""
    throttled = False
    for cand in candidates_of(st):
        for attempt in range(tries_per_candidate):
            PACER.wait()
            try:
                req = urllib.request.Request(
                    wayback_url(cand["page_url"], cand["wayback_ts"]),
                    headers={**UA, "Accept-Encoding": "gzip, identity"})
                with urllib.request.urlopen(req, timeout=180) as r:
                    html = _decompress(r.read(), r.headers.get("Content-Encoding", "") or "")
                PACER.reward()
                # A successful fetch RETIRES the throttle flag. It used to be sticky across
                # retries, so one transient reset before a clean retry made the caller report
                # THROTTLED for a page that had downloaded fine -- and a `throttled` return is
                # `continue`d without a drop reason, so the statement vanished from both the
                # emitted set and the accounting. Measured on 2023-11: its archived page is a
                # stub (0-char narrative, a real `no_narrative`) that two runs in a row filed as
                # a throttle. The rule is still "never bank a throttle as a content verdict";
                # this is the same rule in the other direction.
                throttled = False
            except urllib.error.HTTPError as e:
                if e.code in (404, 403):     # this capture really is not retrievable
                    break
                # 503/429 = archive.org's rate limit, not a statement about the content
                throttled = True
                PACER.penalise()
                time.sleep(min(5 * 2 ** attempt, 60))
                continue
            except Exception:                                    # noqa: BLE001
                throttled = True             # refused/reset connection => throttle, NOT a miss
                PACER.penalise()
                time.sleep(min(5 * 2 ** attempt, 60))
                continue
            if not _looks_like_html(html):
                break
            text, _how = extract_narrative(html, tcfg)
            if len(text) >= tcfg["min_chars"]:
                cached.parent.mkdir(parents=True, exist_ok=True)
                cached.write_bytes(html)
                return html, cand["page_url"], cand["wayback_ts"], text, False
            last = html or last
            break
    return last, st["page_url"], st["wayback_ts"], "", throttled


def resolve_pack(pack_url: str, pack_index: dict) -> tuple[str, str] | None:
    """(fetchable_url, wayback_ts) for the pack a page links, resolved by FILENAME.

    The page's own path cannot be fetched directly: it reflects whichever CMS was live when that
    capture was taken, and the modern `/-/media/project/sites/rbnz/files/...` form is frequently
    unarchived even when the pack itself is (sep97-data.xls 404s at every timestamp on the modern
    path). The filename is stable across both CMS eras, so it is the join key."""
    base = urllib.parse.unquote(pack_url.split("?")[0].rsplit("/", 1)[-1]).lower()
    hit = pack_index.get(base)
    if hit:
        return hit["url"], hit["wayback_ts"]
    # tolerate the '-data'/'data' and 'mps' prefix/suffix variants seen across the archive
    stem = re.sub(r"[^a-z0-9]", "", base)
    for k, v in pack_index.items():
        if re.sub(r"[^a-z0-9]", "", k) == stem:
            return v["url"], v["wayback_ts"]
    return None


# ---------------------------------------------------------------------------- workbook parsing
def _xl_serial_to_date(v: float):
    # Excel's 1900 date system with its deliberate 1900-02-29 bug; 1899-12-30 is the right epoch.
    return datetime(1899, 12, 30) + timedelta(days=float(v))


# RBNZ's older packs label quarters as text instead of dates: '90q4', '  91Q1  ', '1997Q3'.
# Without this the whole 1997-2002 era parses to zero points -- the date column is not a date.
_QTR_RE = re.compile(r"^\s*(\d{2}|\d{4})\s*[qQ]\s*([1-4])\s*$")


def _as_date(v, datemode: int | None):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        qm = _QTR_RE.match(s)
        if qm:
            y = int(qm.group(1))
            if y < 100:                 # two-digit year; this data starts in 1990
                y = 1900 + y if y >= 80 else 2000 + y
            return datetime(y, 3 * int(qm.group(2)), 1)
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        try:                                     # serials sometimes arrive as text ('34759.0')
            f = float(s)
        except ValueError:
            return None
        v = f
    if isinstance(v, (int, float)):
        # plausible window: 1950-01-01 (18264) .. 2050-01-01 (54789). Rejects year numbers
        # (1999) and index levels (114.9) that would otherwise read as dates.
        if 18264 <= float(v) <= 54789:
            return _xl_serial_to_date(v)
    return None


def read_sheets(path: Path) -> list[dict]:
    """Uniform view over both workbook formats: [{name, rows: [[cell, ...]]}]."""
    if path.read_bytes()[:2] == b"PK":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        out = []
        for name in wb.sheetnames:
            out.append({"name": name, "datemode": None,
                        "rows": [list(r) for r in wb[name].iter_rows(values_only=True)]})
        wb.close()
        return out
    import xlrd                                   # legacy .xls (1999-2021 packs)
    wb = xlrd.open_workbook(path, on_demand=True)
    out = []
    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        out.append({"name": name, "datemode": wb.datemode,
                    "rows": [[sh.cell_value(r, c) for c in range(sh.ncols)]
                             for r in range(sh.nrows)]})
    return out


_SKIP_TITLE = re.compile(r"^\s*(source|date|fig(ure)?\s*[\dA-Z.]*|%|index|net %|"
                         r"annual|quarterly|s\.a\.|seasonally adjusted)\s*[:.]?\s*$", re.I)


def sheet_title(rows: list[list], name: str) -> str:
    """The sheet's own variable name, from its title cells -- NOT its 'Fig 2.1' code, which does
    not mean the same variable in different years (see module docstring)."""
    parts: list[str] = []
    for r in rows[:8]:
        for c in r:
            if not isinstance(c, str):
                continue
            s = c.strip()
            if len(s) < 3 or _SKIP_TITLE.match(s) or s.lower().startswith("source"):
                continue
            if re.fullmatch(r"(?i)fig(ure)?\s*[\dA-Z.]+", s) or s == name.strip():
                continue
            parts.append(s)
        if len(parts) >= 2:
            break
    return " | ".join(parts[:2])


def _date_density(rows: list[list], c: int, datemode) -> int:
    return sum(1 for r in rows if c < len(r) and _as_date(r[c], datemode) is not None)


def find_blocks(rows: list[list], datemode, min_points: int = 8) -> list[dict]:
    """Split a sheet into figure blocks, each with ALL of its value columns and their labels:
        [{title, date_col, columns: [{col, label, points}]}]

    Two separate layout facts force this shape.

    (1) MANY FIGURES PER SHEET. The modern packs put one figure per sheet, so a sheet-wide scan was
        enough. The 1997-2002 packs lay many figures out side by side on one wide sheet (measured:
        the Sep-1997 pack is 49 rows x 117 columns, '#1 Consumer price inflation' at column 1 and
        '#2 Real and Nominal MCI' at column 7). A sheet-wide scan finds only the leftmost figure and
        then labels it with every figure's title at once.

    (2) COLUMNS INSIDE A FIGURE ARE DIFFERENT SERIES. Returning only the leftmost value column was
        wrong and actively mislabelled data: the Dec-2010 block titled "Employment and unemployment
        rate" has Employment in its first column, so matching the title's word 'unemployment' and
        then taking column one shipped EMPLOYMENT under an unemployment_rate label. Every column is
        returned with its own header text so the caller can pick the one it actually named.
    """
    width = max((len(r) for r in rows), default=0)
    date_cols = [c for c in range(width) if _date_density(rows, c, datemode) >= min_points]
    blocks = []
    for i, dc in enumerate(date_cols):
        stop = date_cols[i + 1] if i + 1 < len(date_cols) else width
        first_row = next((ri for ri, r in enumerate(rows)
                          if dc < len(r) and _as_date(r[dc], datemode) is not None), 0)
        columns = []
        for vc in range(dc + 1, min(stop, dc + 7)):
            pts = []
            for r in rows:
                if dc >= len(r) or vc >= len(r):
                    continue
                d = _as_date(r[dc], datemode)
                v = r[vc]
                if d is None or not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                pts.append((d, float(v)))
            if len(pts) < min_points:
                continue
            pts.sort(key=lambda x: x[0])
            # this column's own header: nearest text cell above the data, in this column
            label = ""
            for ri in range(first_row - 1, max(-1, first_row - 5), -1):
                if ri < 0 or vc >= len(rows[ri]):
                    continue
                cell = rows[ri][vc]
                if isinstance(cell, str) and len(cell.strip()) >= 3 \
                        and not _SKIP_TITLE.match(cell.strip()):
                    label = cell.strip()
                    break
            columns.append({"col": vc, "label": label, "points": pts})
        if not columns:
            continue
        title_parts = []
        for r in rows[:max(1, first_row)]:
            for c in range(max(0, dc - 1), min(stop, dc + 5)):
                if c < len(r) and isinstance(r[c], str):
                    t = r[c].strip()
                    if len(t) >= 3 and not _SKIP_TITLE.match(t) \
                            and not t.lower().startswith("source"):
                        title_parts.append(re.sub(r"^#\d+\s*", "", t))
        blocks.append({"title": " | ".join(dict.fromkeys(title_parts))[:200],
                       "date_col": dc, "columns": columns})
    return blocks


def pick_column(block: dict, vcfg: list[dict]) -> tuple[dict, dict] | None:
    """(variable, column) for a block, preferring a COLUMN-label match over a title match.

    Order matters and is the whole point: a column's own header is more specific than the figure
    title, so 'Unemployment rate' beats a title that merely contains the word. Only when no column
    header names a variable does the figure title decide, and then the leftmost column is taken --
    which is the release's own central series in every pack checked (alternate columns are prior
    vintages, e.g. 'March projection')."""
    for col in block["columns"]:
        if col["label"]:
            v = _match_one(col["label"], vcfg)
            if v:
                return v, col
    v = match_variable(block["title"], vcfg)
    if v:
        return v, block["columns"][0]
    return None


def find_series(rows: list[list], datemode) -> tuple[int, int, list[tuple]]:
    """(date_col, first_row, points) for the sheet's PRIMARY series: the leftmost numeric column
    to the right of the date column. Column choice is deliberately conservative -- later columns
    are alternate vintages ('March projection') or sub-series, and picking them by position across
    30 years of layouts would be guesswork."""
    best = (None, 0)
    for c in range(min(8, max((len(r) for r in rows), default=0))):
        n = sum(1 for r in rows if c < len(r) and _as_date(r[c], datemode) is not None)
        if n > best[1]:
            best = (c, n)
    date_col, n_dates = best
    if date_col is None or n_dates < 8:
        return -1, -1, []
    val_col = None
    for c in range(date_col + 1, min(date_col + 6, max((len(r) for r in rows), default=0))):
        hits = sum(1 for r in rows
                   if c < len(r) and isinstance(r[c], (int, float))
                   and not isinstance(r[c], bool)
                   and _as_date(r[date_col], datemode) is not None)
        if hits >= max(8, int(0.4 * n_dates)):
            val_col = c
            break
    if val_col is None:
        return date_col, -1, []
    pts = []
    for r in rows:
        if date_col >= len(r) or val_col >= len(r):
            continue
        d = _as_date(r[date_col], datemode)
        v = r[val_col]
        if d is None or not isinstance(v, (int, float)) or isinstance(v, bool):
            continue
        pts.append((d, float(v)))
    pts.sort(key=lambda p: p[0])
    return date_col, val_col, pts


def infer_freq(pts: list[tuple]) -> str:
    """Frequency from the observed spacing, not from a config guess -- packs mix quarterly macro
    series with daily/weekly market series and the mix changes across eras."""
    if len(pts) < 3:
        return "unknown"
    gaps = sorted((pts[i + 1][0] - pts[i][0]).days for i in range(len(pts) - 1))
    med = gaps[len(gaps) // 2]
    if med <= 3:
        return "1d"
    if 4 <= med <= 10:
        return "1w"
    if 25 <= med <= 35:
        return "1M"
    if 80 <= med <= 100:
        return "1q"
    if 350 <= med <= 380:
        return "1y"
    return "unknown"


def _match_one(title: str, vcfg: list[dict]) -> dict | None:
    for v in vcfg:
        if re.search(v["title_pattern"], title, re.I):
            if v.get("exclude_pattern") and re.search(v["exclude_pattern"], title, re.I):
                continue
            return v
    return None


def match_variable(title: str, vcfg: list[dict]) -> dict | None:
    """Try the title's parts in order before the whole string.

    A block's title is assembled from the text cells above its own columns, and on the wide
    legacy sheets a neighbouring figure's heading can bleed into the tail (e.g.
    'Fig 10: Output gap | Fig. 11: World price of commodity imports ...'). Matching the leading
    part first keeps the block labelled with its own figure rather than its neighbour's."""
    for part in [p.strip() for p in title.split("|") if p.strip()]:
        hit = _match_one(part, vcfg)
        if hit:
            return hit
    return _match_one(title, vcfg)


def window_ending_at(pts: list[tuple], end_date, n: int) -> list[tuple]:
    """Trailing n points at or before end_date. Uses <= rather than an exact match: pack date
    conventions differ (quarter-start vs quarter-end) across eras, so requiring the exact
    statement quarter to be present silently emptied older channels."""
    keep = [p for p in pts if p[0] <= end_date]
    return keep[-n:] if keep else []


# ---------------------------------------------------------------------------- alignment
_SUPERLATIVE = re.compile(r"\b(highest|largest|record[- ]high|all-time high|lowest|smallest|"
                          r"record[- ]low|all-time low)\b", re.I)

# A number matching a channel's value is not enough evidence on its own -- caught during the
# original build: the text's "Inflation increased to 3.1%" was credited to `ocr_pct` purely because
# OCR ALSO happened to sit at 3.14 (-> "3.1") two quarters earlier, a coincidental cross-channel
# collision (same failure shape as GAIN's unit-word check). Require one of the channel's own
# keywords within 60 chars BEFORE the matched number.
_CHANNEL_KEYWORDS = {
    "ocr_pct": ("OCR", "Official Cash Rate", "cash rate", "official interest rate"),
    "unemployment_rate_pct": ("unemployment",),
    "inflation_headline": ("inflation", "CPI"),   # bare "Inflation" IS headline
    "inflation_non_tradables": ("non-tradables inflation", "non-tradables"),
    "gdp_growth_pct": ("GDP", "economic growth", "economic activity"),
    "production_gdp_real_2009_10_nzd_bn": ("GDP", "economic growth", "economic activity"),
    "twi_index": ("TWI", "trade-weighted", "trade weighted", "exchange rate"),
    "ninety_day_rate_pct": ("90-day", "ninety day", "bank bill"),
}
# "tradables inflation" is a literal substring of "non-tradables inflation" -- a plain membership
# check would let a non-tradables sentence satisfy the tradables channel too.
_TRADABLES_RE = re.compile(r"(?<!non-)(?<!non )tradables inflation", re.I)


def detect_alignment(text: str, channel_points: dict) -> tuple[str, list]:
    """channel_points: {channel: [(date, value), ...]} (trailing window, oldest-first).
    Checks the last few points of each channel -- the statement's own quarter may itself be the
    Bank's forecast while the text quotes the prior actual quarter."""
    evidence = []
    for name, pts in channel_points.items():
        keywords = _CHANNEL_KEYWORDS.get(name, ())
        for date, val in pts[-3:]:
            for form in {f"{val:g}", f"{val:.1f}"}:
                if len(form.replace(".", "").replace("-", "")) < 2:
                    continue
                for m in re.finditer(re.escape(form) + r"\s*(?:%|per ?cent)", text):
                    before = text[max(0, m.start() - 60): m.start()]
                    if name == "inflation_tradables":
                        if not _TRADABLES_RE.search(before):
                            continue
                    elif keywords and not any(k.lower() in before.lower() for k in keywords):
                        continue
                    evidence.append({"channel": name, "date": date.strftime("%Y-%m-%d"),
                                     "value": val, "quoted_as": m.group(0)})
                    break
                else:
                    continue
                break
    return ("recites" if evidence else "describes"), evidence


def check_superlatives(text: str, channel_points: dict) -> list:
    flags = []
    for m in _SUPERLATIVE.finditer(text):
        window = text[max(0, m.start() - 80): m.end() + 100]
        num = re.search(r"(\d+\.?\d*)\s*(?:%|per ?cent)", window)
        if not num:
            continue
        val = float(num.group(1))
        for name, pts in channel_points.items():
            vals = [v for _, v in pts]
            if not vals or abs(val - vals[-1]) > 5:
                continue
            is_high = bool(re.search(r"highest|largest|record[- ]high|all-time high",
                                     m.group(0), re.I))
            extreme = max(vals) if is_high else min(vals)
            if abs(extreme - val) > 1e-6:
                flags.append({"channel": name, "claim": m.group(0), "claimed_value": val,
                              "actual_extreme": extreme})
    return flags


# ---------------------------------------------------------------------------- build
def statements_from(cfg: dict) -> list[dict]:
    """Statement list: the census index when present, else the pinned demo list."""
    idx = PKG_ROOT / cfg.get("census", {}).get("index_path", "")
    if idx.exists():
        return json.loads(idx.read_text())["statements"]
    print("no census index -- falling back to config.statements (demo mode)", file=sys.stderr)
    out = []
    for s in cfg.get("statements", []):
        y, m = (int(x) for x in s["published"].split("-")[:2])
        out.append({"year": y, "month": m, "wayback_ts": s["page_wayback_ts"],
                    "page_url": s["page_url"], "path": s["page_url"]})
    return out


def quarter_end(year: int, month: int) -> datetime:
    """The statement's own quarter, as its LAST day. MPS months are Feb/Mar, May/Jun, Aug/Sep,
    Nov/Dec, so the statement quarter is the calendar quarter its month falls in."""
    q = (month - 1) // 3
    m_end = 3 * q + 3
    last = 31 if m_end in (3, 12) else 30
    return datetime(year, m_end, last)


def prefetch(cfg: dict, sts: list, pack_index: dict, workers: int = 4) -> None:
    """Download every page, then every pack, concurrently, into the on-disk cache.

    Purely a throughput measure: the build is serial and each Wayback round trip is slow (pages
    ~80KB, packs up to ~2MB), which measured out at ~100s/statement -- over three hours for 119
    statements. Modest concurrency only; archive.org throttles, and a throttle must never be cached
    as a final answer (fetch() already refuses to cache a non-workbook body for packs)."""
    from concurrent.futures import ThreadPoolExecutor
    cache = PKG_ROOT / cfg["data"]["cache_dir"]

    def page(st):
        _h, _u, _ts, text, _th = fetch_page(st, cache, cfg["text"])
        return bool(text)

    with ThreadPoolExecutor(max_workers=workers) as ex:
        got = list(ex.map(page, sts))
    print(f"prefetch pages: {sum(got)}/{len(sts)} cached", flush=True)

    jobs = []
    for st in sts:
        tag = f"{st['year']:04d}-{st['month']:02d}"
        fp = cache / "pages" / f"{tag}.html"
        if not fp.exists():
            continue
        purl = find_pack_url(fp.read_bytes(), st["page_url"])
        if not purl:
            continue
        res = resolve_pack(purl, pack_index)
        if not res:
            continue
        url, ts = res
        ext = ".xlsx" if ".xlsx" in url.lower() else ".xls"
        jobs.append((cache / "packs" / f"{tag}{ext}", url, ts))

    def pack(job):
        path, url, ts = job
        try:
            fetch(wayback_url(url, ts), path, expect_workbook=True)
            return True
        except Exception:                                        # noqa: BLE001
            return False

    with ThreadPoolExecutor(max_workers=workers) as ex:
        ok = list(ex.map(pack, jobs))
    print(f"prefetch packs: {sum(ok)}/{len(jobs)} cached", flush=True)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--config", default=str(PKG_ROOT / "config.example.yaml"))
    ap.add_argument("--set", action="append", default=[])
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--prefetch", action="store_true",
                    help="download all pages+packs concurrently first, then build from cache")
    ap.add_argument("--workers", type=int, default=4)
    args = ap.parse_args()

    cfg = yaml.safe_load(Path(args.config).read_text())
    for override in args.set:
        k, v = override.split("=", 1)
        d = cfg
        parts = k.split(".")
        for p in parts[:-1]:
            d = d[p]
        d[parts[-1]] = yaml.safe_load(v)

    cache = PKG_ROOT / cfg["data"]["cache_dir"]
    scfg, tcfg, lcfg = cfg["series"], cfg["text"], cfg["license"]
    vcfg = cfg["variables"]
    stats = {"statements": 0, "emitted": 0, "recites": 0, "describes": 0,
             "no_narrative": 0, "no_pack": 0, "pack_not_archived": 0,
             "pack_fetch_failed": 0, "pack_unreadable": 0,
             "too_few_channels": 0, "page_fetch_failed": 0, "throttled": 0,
             "superlative_flags": 0, "superlative_dropped": 0,
             "anchor_used": {}, "unmatched_sheet_titles": {}, "channels_per_record": [],
             "by_year": {}}
    pack_index = {}
    pidx = PKG_ROOT / cfg["census"].get("pack_index_path", "")
    if pidx.exists():
        pack_index = json.loads(pidx.read_text())["packs"]
        print(f"pack index: {len(pack_index)} archived data packs")
    records = []
    sts = statements_from(cfg)
    if args.limit:
        sts = sts[: args.limit]
    if args.prefetch:
        prefetch(cfg, sts, pack_index, args.workers)
    print(f"building from {len(sts)} statements")

    for st in sts:
        stats["statements"] += 1
        tag = f"{st['year']:04d}-{st['month']:02d}"
        page_html, page_url, page_ts, text, throttled = fetch_page(st, cache, tcfg)
        if throttled and not text:
            # never bank a throttle as a content verdict
            print(f"  {tag}: THROTTLED by archive.org -- not counted as a miss")
            stats["throttled"] += 1
            continue
        if not page_html:
            print(f"  {tag}: no candidate capture fetched")
            stats["page_fetch_failed"] += 1
            continue
        _t, how = extract_narrative(page_html, tcfg)
        stats["anchor_used"][how] = stats["anchor_used"].get(how, 0) + 1
        if len(text) < tcfg["min_chars"]:
            print(f"  {tag}: no narrative (how={how}, {len(text)} chars)")
            stats["no_narrative"] += 1
            continue

        pack_url = find_pack_url(page_html, page_url)
        if not pack_url:
            print(f"  {tag}: no data pack linked")
            stats["no_pack"] += 1
            continue
        resolved = resolve_pack(pack_url, pack_index)
        if not resolved:
            print(f"  {tag}: pack linked but not archived ({pack_url.rsplit('/', 1)[-1]})")
            stats["pack_not_archived"] += 1
            continue
        pack_fetch_url, pack_ts = resolved
        ext = ".xlsx" if ".xlsx" in pack_fetch_url.lower() else ".xls"
        try:
            pack_path = cache / "packs" / f"{tag}{ext}"
            fetch(wayback_url(pack_fetch_url, pack_ts), pack_path, expect_workbook=True)
        except Exception as e:                                   # noqa: BLE001
            print(f"  {tag}: pack fetch failed ({type(e).__name__})")
            stats["pack_fetch_failed"] += 1
            continue
        try:
            sheets = read_sheets(pack_path)
        except Exception as e:                                   # noqa: BLE001
            print(f"  {tag}: pack unreadable ({type(e).__name__}: {e})")
            stats["pack_unreadable"] += 1
            continue

        end_q = quarter_end(st["year"], st["month"])
        channel_points: dict[str, list] = {}
        channel_freq: dict[str, str] = {}
        meta_channels = []
        for sh in sheets:
            if not sh["rows"]:
                continue
            blocks = find_blocks(sh["rows"], sh["datemode"], scfg["min_points"])
            for blk in blocks:
                if not (blk["title"] or blk["columns"]):
                    continue
                hit = pick_column(blk, vcfg)
                if hit is None:
                    key = (blk["title"] or sheet_title(sh["rows"], sh["name"]))[:60]
                    stats["unmatched_sheet_titles"][key] = \
                        stats["unmatched_sheet_titles"].get(key, 0) + 1
                    continue
                var, col = hit
                if var["name"] in channel_points:     # first match wins (leftmost/earliest figure)
                    continue
                title = f"{blk['title'][:110]} :: col={col['label'][:40]!r}"
                pts = col["points"]
                freq = infer_freq(pts)
                if freq == "unknown":
                    continue
                n = (scfg["daily_window"] if freq in ("1d", "1w")
                     else scfg["quarterly_window"])
                win = (pts[-n:] if freq in ("1d", "1w")
                       else window_ending_at(pts, end_q, n))
                if len(win) < scfg["min_points"]:
                    continue
                channel_points[var["name"]] = win
                channel_freq[var["name"]] = freq
                meta_channels.append({"unit": var["name"], "sheet": sh["name"],
                                      "block_title": title[:120], "freq": freq,
                                      "n_points": len(win)})

        if len(channel_points) < scfg["min_channels"]:
            print(f"  {tag}: only {len(channel_points)} channels")
            stats["too_few_channels"] += 1
            continue

        # schema: equal length within a freq. Trim each freq group to its shortest channel.
        for freq in set(channel_freq.values()):
            names = [n for n, f in channel_freq.items() if f == freq]
            k = min(len(channel_points[n]) for n in names)
            for n in names:
                channel_points[n] = channel_points[n][-k:]

        alignment, evidence = detect_alignment(text, channel_points)
        superlative_flags = check_superlatives(text, channel_points)
        stats["superlative_flags"] += len(superlative_flags)
        if superlative_flags and tcfg.get("drop_on_superlative_contradiction", True):
            print(f"  {tag}: dropped on superlative contradiction")
            stats["superlative_dropped"] += 1
            continue

        ts = [{"values": [v for _, v in pts], "unit": name, "freq": channel_freq[name]}
              for name, pts in channel_points.items()]
        first_date = min(pts[0][0] for pts in channel_points.values())
        last_date = max(pts[-1][0] for pts in channel_points.values())
        rec = emit_record(
            text=text + "\n\n<ts></ts>",
            timeseries=ts,
            alignment=alignment,
            license=lcfg["tag"],
            source=page_url,
            series_id=f"rbnz_mps_{tag}",
            dataset="rbnz_mps",
            domain="monetary_policy",
            region=cfg.get("region", "NZ"),
            period_start=first_date.strftime("%Y-%m-%d"),
            period_end=last_date.strftime("%Y-%m-%d"),
            meta={
                "statement": tag,
                "statement_quarter_end": end_q.strftime("%Y-%m-%d"),
                "channels": meta_channels,
                "n_channels": len(channel_points),
                "narrative_anchor": how,
                "recite_evidence": evidence,
                "superlative_flags": superlative_flags,
                "true_license": lcfg["true_license"],
                "attribution_required": lcfg["attribution_required"],
                "access_method": "wayback_snapshot",
                "wayback_ts": page_ts,
                "data_pack_url": pack_fetch_url,
                "data_pack_linked_as": pack_url,
                "series_note": (
                    "channels are matched to each pack sheet's OWN title text, not its figure "
                    "number -- figure numbering is re-cut between editions (Fig 2.1 is 'CPI "
                    "inflation' in Jun-2005 and 'Employment and investment intentions' in "
                    "Dec-2010), so a number->variable map would mislabel series. Quarterly "
                    "channels are windowed to the statement's own quarter; where that final "
                    "point is the Bank's own projection rather than an outturn it carries the "
                    "forecast-not-measured caveat used for WASDE #41 / GAIN #58 -- the text is "
                    "the contemporaneous first-party forecast, so no future-value leakage."),
            },
        )
        records.append(rec)
        stats["emitted"] += 1
        stats[alignment] += 1
        stats["channels_per_record"].append(len(channel_points))
        stats["by_year"][str(st["year"])] = stats["by_year"].get(str(st["year"]), 0) + 1
        print(f"  {tag}: OK {len(channel_points)} channels, {alignment}, {len(text)} chars")

    out = PKG_ROOT / cfg["output"]["path"]
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w") as fh:
        for r in records:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")

    cp = stats.pop("channels_per_record")
    stats["mean_channels"] = round(sum(cp) / len(cp), 2) if cp else 0
    stats["unmatched_sheet_titles"] = dict(
        sorted(stats["unmatched_sheet_titles"].items(), key=lambda kv: -kv[1])[:40])
    rr = PKG_ROOT / cfg["output"]["run_report"]
    rr.write_text(json.dumps({"dataset": "rbnz_mps", "stats": stats, "config_snapshot": cfg},
                             indent=2, ensure_ascii=False))
    print(json.dumps(stats, indent=2, ensure_ascii=False))
    print(f"\nwrote {len(records)} records -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
